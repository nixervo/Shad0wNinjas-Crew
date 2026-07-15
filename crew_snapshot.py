import os
import urllib.request
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import json
import time
import sys
import base64
import re

CREW_ID = 101  # Shad0w Ninjas

API_URL = f"https://playninjarift.com/api/detail_crew_website.php?crew_id={CREW_ID}"
TARGET_TZ = timezone(timedelta(hours=8))
EXCEL_FILE = f"crew_{CREW_ID}.xlsx"
HOURLY_CACHE = "_hourly_cache.json"
CACHE_30M = "_30m_cache.json"
CACHE_30M_KILLS = "_30m_kills_cache.json"
CACHE_1H = "_1h_cache.json"
CACHE_1H_KILLS = "_1h_kills_cache.json"
CHANGES_JSON = "_changes.json"
PHASE1_CACHE = "_phase1_cache.json"
CASTLE_API = "https://playninjarift.com/api/crew_ranking_castles_website.php"
CASTLE_CACHE_30M = "_30m_castle_cache.json"

def fetch_crew():
    req = urllib.request.Request(API_URL, headers={"User-Agent": "Crew-snapshot/1.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode())

SEASON_API = "https://playninjarift.com/api/refresh_time_crew_website.php"
RANKING_API = "https://playninjarift.com/api/crew_ranking_website.php"

def fetch_season_info():
    req = urllib.request.Request(SEASON_API, headers={"User-Agent": "Crew-snapshot/1.0"})
    with urllib.request.urlopen(req, timeout=10) as resp:
        return json.loads(resp.read().decode())

def fetch_crew_ranking():
    req = urllib.request.Request(RANKING_API, headers={"User-Agent": "Crew-snapshot/1.0"})
    with urllib.request.urlopen(req, timeout=10) as resp:
        data = json.loads(resp.read().decode())
    for entry in data:
        if entry["crew_id"] == CREW_ID:
            return entry
    return {}

def fetch_castle_ranking():
    req = urllib.request.Request(CASTLE_API, headers={"User-Agent": "Crew-snapshot/1.0"})
    with urllib.request.urlopen(req, timeout=10) as resp:
        return json.loads(resp.read().decode())

def get_previous_sheet_names(wb):
    names = [s.title for s in wb.worksheets]
    names.sort()
    return names

def get_unique_names(members):
    """Returns list of (api_name, display_name) tuples."""
    name_count = {}
    for m in members:
        nm = m["character_name"]
        name_count[nm] = name_count.get(nm, 0) + 1
    counter = {}
    result = []
    for m in members:
        nm = m["character_name"]
        if name_count[nm] > 1:
            counter[nm] = counter.get(nm, 0) + 1
            result.append((nm, f"{nm} (#{counter[nm]})"))
        else:
            result.append((nm, nm))
    return result

def load_prev_from_xlsx(filename, before_date):
    prev_data = []
    prev_timestamp = None
    if not os.path.exists(filename):
        return prev_data, prev_timestamp
    wb = load_workbook(filename)
    prev_names = [s.title for s in wb.worksheets]
    prev_names.sort()
    prev_sheet_name = None
    for n in reversed(prev_names):
        if n < before_date:
            prev_sheet_name = n
            break
    if prev_sheet_name and prev_sheet_name in wb.sheetnames:
        ps = wb[prev_sheet_name]
        raw = ps["A2"].value
        if raw:
            prev_timestamp = raw.replace("Timestamp: ", "")
        for row in ps.iter_rows(min_row=4, max_col=2, values_only=True):
            if row[0] and row[1] is not None:
                prev_data.append({"character_name": row[0], "member_damage": int(row[1])})
    return prev_data, prev_timestamp

def load_hourly_cache():
    if not os.path.exists(HOURLY_CACHE):
        return {}, None
    with open(HOURLY_CACHE, encoding="utf-8") as f:
        c = json.load(f)
    return c.get("members", {}), c.get("timestamp")

def save_hourly_cache(members, unique_names, now):
    cache = {
        "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
        "members": {unique_names[i][1]: m["member_damage"] for i, m in enumerate(members)},
    }
    with open(HOURLY_CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f)

def load_30m_cache():
    if not os.path.exists(CACHE_30M):
        return None
    with open(CACHE_30M, encoding="utf-8") as f:
        return json.load(f)

def save_30m_cache(members, unique_names):
    cache = {
        "timestamp": datetime.now(TARGET_TZ).strftime("%Y-%m-%d %H:%M:%S"),
        "members": {unique_names[i][1]: m["member_damage"] for i, m in enumerate(members)},
        "order": [unique_names[i][1] for i, _ in enumerate(members)],
    }
    with open(CACHE_30M, "w", encoding="utf-8") as f:
        json.dump(cache, f)

def load_30m_kills_cache():
    if not os.path.exists(CACHE_30M_KILLS):
        return None
    with open(CACHE_30M_KILLS, encoding="utf-8") as f:
        return json.load(f)

def save_30m_kills_cache(members, unique_names):
    cache = {
        "timestamp": datetime.now(TARGET_TZ).strftime("%Y-%m-%d %H:%M:%S"),
        "members": {unique_names[i][1]: (m.get("boss_kills", 0) or 0) for i, m in enumerate(members)},
    }
    with open(CACHE_30M_KILLS, "w", encoding="utf-8") as f:
        json.dump(cache, f)

def load_1h_cache():
    if not os.path.exists(CACHE_1H):
        return None
    with open(CACHE_1H, encoding="utf-8") as f:
        return json.load(f)

def save_1h_cache(member_dict, timestamp):
    with open(CACHE_1H, "w", encoding="utf-8") as f:
        json.dump({"timestamp": timestamp, "members": member_dict}, f)

def load_1h_kills_cache():
    if not os.path.exists(CACHE_1H_KILLS):
        return None
    with open(CACHE_1H_KILLS, encoding="utf-8") as f:
        return json.load(f)

def save_1h_kills_cache(member_dict, timestamp):
    with open(CACHE_1H_KILLS, "w", encoding="utf-8") as f:
        json.dump({"timestamp": timestamp, "members": member_dict}, f)

def load_30m_castle_cache():
    if not os.path.exists(CASTLE_CACHE_30M):
        return None
    with open(CASTLE_CACHE_30M, encoding="utf-8") as f:
        return json.load(f)

def save_30m_castle_cache(castle_data, timestamp):
    with open(CASTLE_CACHE_30M, "w", encoding="utf-8") as f:
        json.dump({"timestamp": timestamp, "castles": castle_data}, f)

def compute_rolling_avg_daily_gain(filename, before_date):
    if not os.path.exists(filename):
        return None
    wb = load_workbook(filename)
    names = sorted([s.title for s in wb.worksheets if s.title != "Sheet1" and s.title < before_date])
    gains = []
    for i in range(1, len(names)):
        prev_total = 0
        curr_total = 0
        ps = wb[names[i-1]]
        cs = wb[names[i]]
        for row in ps.iter_rows(min_row=4, max_col=2, values_only=True):
            if row[1] is not None:
                prev_total += int(row[1])
        for row in cs.iter_rows(min_row=4, max_col=2, values_only=True):
            if row[1] is not None:
                curr_total += int(row[1])
        gains.append(curr_total - prev_total)
    if not gains:
        return None
    return sum(gains) / len(gains)

def compute_changes(members, prev_data):
    prev_names = {m["character_name"] for m in prev_data}
    today_names = {m["character_name"] for m in members}
    left_names = sorted(prev_names - today_names)
    joined_names = sorted(today_names - prev_names)
    return left_names, joined_names

def compute_diff(members, prev_data):
    prev_map = {m["character_name"]: m["member_damage"] for m in prev_data}
    result = []
    for m in members:
        name = m["character_name"]
        reps = m["member_damage"]
        if name in prev_map:
            diff = reps - prev_map[name]
            diff_str = f"+{diff}" if diff > 0 else str(diff)
        else:
            diff_str = "N/A"
        result.append((name, reps, diff_str))
    return result

def write_sheet(ws, data, prev_data, now, unique_names):
    uniq = unique_names if unique_names else get_unique_names(data["members"])
    rows = compute_diff(data["members"], prev_data)
    daily_lookup = {name: diff for name, _, diff in rows}
    names = [uniq[i][1] for i in range(len(uniq))]
    reps = [str(m["member_damage"]) for m in data["members"]]
    diffs = [daily_lookup.get(m["character_name"], "N/A") for m in data["members"]]
    ws.title = now.strftime("%Y-%m-%d")

    max_name = max((len(n) for n in names), default=10)
    max_reps = max((len(r) for r in reps), default=4)
    max_diff = max((len(d) for d in diffs), default=3)

    ws.column_dimensions["A"].width = max(max_name + 5, 10)
    ws.column_dimensions["B"].width = max(max_reps + 3, 8)
    ws.column_dimensions["C"].width = max(max_diff + 3, 14)
    ws.column_dimensions["D"].width = max(max_reps + 3, 8)
    ws.column_dimensions["E"].width = max(max_diff + 3, 14)

    crew_name = data.get("crew_name", "Unknown")
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Crew: {crew_name} ({CREW_ID})"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Timestamp: {now.strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Name", "Dmg", "Daily Dmg", "Boss Kills", "Daily Kills"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    kills_map = {m["character_name"]: (m.get("boss_kills", 0) or 0) for m in data["members"]}
    prev_kills_map = {}
    if os.path.exists(EXCEL_FILE):
        pw = load_workbook(EXCEL_FILE)
        pn = sorted([s.title for s in pw.worksheets if s.title != "Sheet1"])
        if len(pn) >= 2:
            ps = pw[pn[-2]]
            for row in ps.iter_rows(min_row=4, max_col=4, values_only=True):
                if row[0] and row[1] is not None:
                    prev_kills_map[row[0]] = int(row[3]) if len(row) >= 4 and row[3] is not None else 0
    for row_idx, (name, reps_val, diff_val) in enumerate(rows, 4):
        ws.cell(row=row_idx, column=1, value=name).alignment = Alignment(vertical="center")
        ws.cell(row=row_idx, column=2, value=reps_val).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=3, value=diff_val).alignment = Alignment(horizontal="center", vertical="center")
        raw_name = uniq[row_idx - 4][0] if uniq else name.split(" (#")[0]
        cur_kills = kills_map.get(raw_name, 0)
        prev_k = prev_kills_map.get(raw_name, cur_kills) if prev_kills_map else cur_kills
        dk = cur_kills - prev_k
        ws.cell(row=row_idx, column=4, value=cur_kills).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=5, value=f"+{dk}" if dk > 0 else str(dk) if prev_kills_map else "N/A").alignment = Alignment(horizontal="center", vertical="center")

def save_xlsx(data, prev_data, now, uniq):
    sheet_name = now.strftime("%Y-%m-%d")

    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)
    write_sheet(ws, data, prev_data, now, uniq)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(EXCEL_FILE)
    print(f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] Saved sheet '{sheet_name}' to {EXCEL_FILE}")

def save_seasonal_xlsx(members, season_num):
    filename = f"S{season_num}_ID{CREW_ID}.xlsx"
    if os.path.exists(filename):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = f"Season {season_num}"
    header = f"[S{season_num}] Total Reps"
    ws["A1"] = "Name"
    ws["B1"] = header
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    max_name = 10
    max_reps = 4
    for idx, m in enumerate(members, 2):
        name = m["character_name"]
        reps = m["member_damage"]
        ws.cell(row=idx, column=1, value=name).alignment = Alignment(vertical="center")
        ws.cell(row=idx, column=2, value=reps).alignment = Alignment(horizontal="center", vertical="center")
        max_name = max(max_name, len(name))
        max_reps = max(max_reps, len(str(reps)))
    ws.column_dimensions["A"].width = max_name + 5
    ws.column_dimensions["B"].width = max_reps + 3
    wb.save(filename)
    print(f"Saved seasonal snapshot: {filename}")

def load_changes():
    try:
        with open(CHANGES_JSON) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []

def save_changes(changes):
    now = datetime.now(TARGET_TZ)
    cutoff = now - timedelta(hours=72)
    pruned = [c for c in changes if c.get("detected_at", "") >= cutoff.strftime("%Y-%m-%d %H:%M:%S")]
    with open(CHANGES_JSON, "w") as f:
        json.dump(pruned, f)

def diff_html(diff_str):
    if diff_str.startswith("+"):
        return f'<span class="up">{diff_str}</span>'
    elif diff_str == "0":
        return f'<span class="down">{diff_str}</span>'
    elif diff_str.startswith("-"):
        return f'<span class="down">{diff_str}</span>'
    else:
        return f'<span class="na">{diff_str}</span>'

def save_html(data, prev_data, prev_timestamp, hourly_diffs, hourly_ts, now, all_dates, show_changes, season_info=None, stats=None, diff_30m=None, changes=None, hourly_cache=None, cache_30m=None, phase1=None, diffs_30m_kills=None, castle_stats=None, diffs_1h_kills=None):
    daily_rows = compute_diff(data["members"], prev_data)
    crew_name = data.get("crew_name", "Unknown")
    date_str = now.strftime("%Y-%m-%d")
    ts_str = now.strftime("%Y-%m-%d %H:%M:%S")
    member_count = len(data["members"])

    logo_b64 = ""
    logo_path = "Crew_logo.png"
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            logo_b64 = base64.b64encode(f.read()).decode()

    favicon_b64 = ""
    if os.path.exists("favicon.ico"):
        with open("favicon.ico", "rb") as f:
            favicon_b64 = base64.b64encode(f.read()).decode()

    archive_links = "".join(
        f'<a href="{d}.html" class="{"active" if d == date_str else ""}">{d}</a>'
        for d in sorted(all_dates, reverse=True)
    )

    diffs_30m_map = diff_30m.get("diffs", {}) if isinstance(diff_30m, dict) else {}
    daily_lookup = {name: diff for name, _, diff in daily_rows}
    kills_30m_map = diffs_30m_kills if diffs_30m_kills else {}
    kills_1h_map = diffs_1h_kills if diffs_1h_kills else {}
    dmg_1h_map = hourly_diffs if hourly_diffs else {}
    uniq_names = get_unique_names(data["members"])
    phase_num = season_info.get("phase", 1) if season_info else 1
    in_p1 = (phase_num == 1)
    p1_colspan = 6 if in_p1 else 2
    p2_colspan = 1 if in_p1 else 3
    hide_css = ".ph2-diff { display: none !important; }" if in_p1 else ".ph1-diff { display: none !important; }"
    na = '<span class="na">N/A</span>'
    table_rows = "".join(
        f"<tr><td class=\"num\">{i+1}</td><td>{uniq_names[i][1]}</td><td class=\"num div-col\">{m.get('boss_kills', 0) or 0:,}</td><td class=\"num ph1-diff\">{diff_html(kills_30m_map.get(uniq_names[i][1], 'N/A'))}</td><td class=\"num ph1-diff\">{diff_html(kills_1h_map.get(uniq_names[i][1], 'N/A'))}</td><td class=\"num\">{m['member_damage']:,}</td><td class=\"num ph1-diff\">{diff_html(diffs_30m_map.get(uniq_names[i][1], 'N/A'))}</td><td class=\"num ph1-diff\">{diff_html(dmg_1h_map.get(uniq_names[i][1], 'N/A'))}</td><td class=\"num div-col\">" + (na if in_p1 else f'{m["member_damage"]:,}') + "</td><td class=\"num ph2-diff\">" + (na if in_p1 else diff_html(diffs_30m_map.get(uniq_names[i][1], 'N/A'))) + "</td><td class=\"num ph2-diff\">" + (na if in_p1 else diff_html(dmg_1h_map.get(uniq_names[i][1], 'N/A'))) + f"</td><td class=\"num div-col\">{diff_html(daily_lookup.get(m['character_name'], 'N/A'))}</td><td class=\"num\">{na}</td></tr>"
        for i, m in enumerate(data["members"])
    )

    changes_html = ""
    if changes:
        left_names = [c["name"] for c in changes if c["type"] == "left"]
        joined_names = [c["name"] for c in changes if c["type"] == "joined"]
        if left_names or joined_names:
            left_items = "".join(f"<li>{n}</li>" for n in left_names)
            joined_items = "".join(f"<li>{n}</li>" for n in joined_names)
            changes_html = f"""
  <div class="changes">
    <div class="changes-title">Member Changes (last 72h)</div>
    <div class="changes-cols">
      <div class="changes-col">
        <div class="changes-head left">Left ({len(left_names)})</div>
        <ul>{left_items}</ul>
      </div>
      <div class="changes-col">
        <div class="changes-head joined">Joined ({len(joined_names)})</div>
        <ul>{joined_items}</ul>
      </div>
    </div>
  </div>"""

    logo_html = f'<img src="data:image/png;base64,{logo_b64}" class="logo" alt="SHAD0W NINJAS">' if logo_b64 else ""
    favicon_html = f'<link rel="icon" type="image/x-icon" href="data:image/x-icon;base64,{favicon_b64}">' if favicon_b64 else ""

    hourly_ref = f"Ref (hourly): {hourly_ts}" if hourly_ts else ""
    ref_30m_val = diff_30m.get("ts", "") if isinstance(diff_30m, dict) else ""
    ref_30m = f"Ref (30m): {ref_30m_val}" if ref_30m_val else ""
    daily_ref = f"Ref (daily): {prev_timestamp}" if prev_timestamp else ""

    timer_html = ""
    season_end_iso = ""
    if season_info:
        season_num = season_info["season"]
        phase_num = season_info.get("phase", 1)
        end_dt = datetime(2026, 7, 19, 5, 0, 0, tzinfo=timezone.utc)
        season_end_iso = "2026-07-19T05:00:00Z"
        timer_html = f"""
  <div class="timer-bar">
    <span class="timer-left">
      <span class="timer-season">Season <span id="season-num">{season_num}</span> &middot; Phase {phase_num}</span>
      <span class="timer-sep">&middot;</span>
      <span class="timer-clock">
        <span class="timer-digits"><span id="timer-d">--</span><span class="timer-unit">d</span></span>
        <span class="timer-digits"><span id="timer-h">--</span><span class="timer-unit">h</span></span>
        <span class="timer-digits"><span id="timer-m">--</span><span class="timer-unit">m</span></span>
        <span class="timer-digits"><span id="timer-s">--</span><span class="timer-unit">s</span></span>
      </span>
    </span>
    <span class="timer-right" id="auto-refresh">&#x21BB; Auto Refresh in <span id="auto-seconds">60</span>s</span>
  </div>"""

    stats_html = ""
    if stats:
        stats_html = f"""
  <div class="stats-bar">
    <div class="stats-row">
      <div class="stats-col">
        <span class="stat-label">Today</span>
        <span class="stat-val" id="today-gain">+{stats['today_gain']:,}</span>
      </div>
      <div class="stats-col">
        <span class="stat-label">Total Damage</span>
        <span class="stat-val" id="season-total">{stats['season_total']:,}</span>
      </div>
    </div>
  </div>"""

    castle_html = ""
    if castle_stats and phase_num == 1:
        castle_cards = ""
        for cs in castle_stats:
            our_g_str = f"+{cs['our_gain']:,}" if cs['our_gain'] > 0 else str(cs['our_gain'])
            rival_g_str = f"+{cs['rival_gain']:,}" if cs['rival_gain'] > 0 else str(cs['rival_gain'])
            is_lead = (cs["rank"] == 1)
            gap_label = "Lead" if is_lead else "Need"
            gap_disp = cs["gap"] if is_lead else cs["rival_kills"] - cs["our_kills"]
            if is_lead:
                top_html = f"""          <div class="castle-emoji">&#127983;</div>
          <div class="castle-name">{cs['name']}</div>
          <div class="castle-rank-pill">[#{cs['rank']}]</div>
          <div class="castle-kills ours">{cs['our_kills']:,}</div>
          <div class="castle-gain ours">({our_g_str}/½h)</div>
          <div class="castle-div"></div>
          <div class="castle-rival-rank">#{cs['rival_rank']}</div>
          <div class="castle-rival-name">{cs['rival_name']}</div>
          <div class="castle-kills rival">{cs['rival_kills']:,}</div>
          <div class="castle-gain rival">({rival_g_str}/½h)</div>"""
            else:
                top_html = f"""          <div class="castle-emoji">&#127983;</div>
          <div class="castle-name">{cs['name']}</div>
          <div class="castle-rival-rank">#{cs['rival_rank']}</div>
          <div class="castle-rival-name">{cs['rival_name']}</div>
          <div class="castle-kills rival">{cs['rival_kills']:,}</div>
          <div class="castle-gain rival">({rival_g_str}/½h)</div>
          <div class="castle-div"></div>
          <div class="castle-rank-pill">[#{cs['rank']}]</div>
          <div class="castle-kills ours">{cs['our_kills']:,}</div>
          <div class="castle-gain ours">({our_g_str}/½h)</div>"""
            castle_cards += f"""        <div class="castle-card" data-castle="{cs['name']}">
{top_html}
          <div class="castle-tag">{gap_label} {gap_disp:,}</div>
        </div>
"""
        castle_html = f"""
  <div class="castle-bar" id="castle-bar">
    <div class="castle-header">CASTLES</div>
    <div class="castle-grid">
{castle_cards}    </div>
  </div>"""

    script_html = ""
    if season_info:
        script_html = """<script>
(function() {
  var end = new Date(\"""" + season_end_iso + """\").getTime();
  function tick() {
    var diff = end - new Date().getTime();
    if (diff <= 0) { document.getElementById("timer-d").textContent = "0"; document.getElementById("timer-h").textContent = "00"; document.getElementById("timer-m").textContent = "00"; document.getElementById("timer-s").textContent = "00"; return; }
    document.getElementById("timer-d").textContent = Math.floor(diff / 86400000);
    document.getElementById("timer-h").textContent = String(Math.floor((diff % 86400000) / 3600000)).padStart(2,"0");
    document.getElementById("timer-m").textContent = String(Math.floor((diff % 3600000) / 60000)).padStart(2,"0");
    document.getElementById("timer-s").textContent = String(Math.floor((diff % 60000) / 1000)).padStart(2,"0");
  }
  tick();
  setInterval(tick, 1000);
})();
window.__crewId = """ + str(CREW_ID) + """;
window.__phase = """ + str(phase_num) + """;
window.__hourlyCache = """ + json.dumps(hourly_cache if hourly_cache else {}) + """;
window.__30mCache = """ + json.dumps(cache_30m["members"] if cache_30m and "members" in cache_30m else {}) + """;
(function() {
  var tbody = document.querySelector("tbody");
  window.__originalRows = tbody.innerHTML;
  window.__defaultRows = tbody.innerHTML;
  var sortCol = -1, sortDir = 0;
  var ths = document.querySelectorAll("th");
  function applySort() {
    if (sortDir === 0) { tbody.innerHTML = window.__originalRows; for (var a = 0; a < ths.length; a++) { var ar = ths[a].querySelector(".sort-arrow"); if (ar) ar.textContent = ""; } if (window.__refreshData) window.__refreshData(); var se = document.getElementById("search-input"); if(se&&se.value){var q=se.value.toLowerCase(),rr=tbody.querySelectorAll("tr");for(var ri=0;ri<rr.length;ri++)rr[ri].style.display=rr[ri].cells[1].textContent.trim().toLowerCase().indexOf(q)>=0?"":"none";} return; }
    for (var a = 0; a < ths.length; a++) { var ar = ths[a].querySelector(".sort-arrow"); if (ar) ar.textContent = ""; }
    for (var a = 0; a < ths.length; a++) {
      if (parseInt(ths[a].getAttribute("data-col")) === sortCol) {
        ths[a].querySelector(".sort-arrow").textContent = sortDir === 1 ? "\\u25B2" : "\\u25BC";
        break;
      }
    }
    var rows = Array.prototype.slice.call(tbody.querySelectorAll("tr"));
    rows.sort(function(a, b) {
      var va = a.cells[sortCol].textContent.trim(), vb = b.cells[sortCol].textContent.trim();
      if (sortCol === 1) return sortDir === 1 ? va.localeCompare(vb) : vb.localeCompare(va);
      var na = parseFloat(va) || -1/0, nb = parseFloat(vb) || -1/0;
      return sortDir === 1 ? na - nb : nb - na;
    });
    for (var r = 0; r < rows.length; r++) tbody.appendChild(rows[r]);
    var sr = tbody.querySelectorAll("tr");
    for (var ri = 0; ri < sr.length; ri++) sr[ri].cells[0].textContent = ri + 1;
  }
  window.__resetSort = function() { sortCol = -1; sortDir = 0; applySort(); };
  for (var i = 0; i < ths.length; i++) (function(col) {
    if (ths[col].hasAttribute("colspan")) return;
    var dcol = parseInt(ths[col].getAttribute("data-col"));
    if (dcol === undefined || isNaN(dcol)) return;
    ths[col].addEventListener("click", function() {
      if (sortCol !== dcol) { sortCol = dcol; sortDir = 1; }
      else { sortDir = (sortDir + 1) % 3; }
      applySort();
      localStorage.setItem("nr_sort", JSON.stringify({col: sortCol, dir: sortDir}));
    });
  })(i);
  try { var _s = JSON.parse(localStorage.getItem("nr_sort")); if (_s && _s.dir > 0) { sortCol = _s.col; sortDir = _s.dir; applySort(); } } catch(e) {}
})();
(function() {
  var API = "https://playninjarift.com/api/detail_crew_website.php?crew_id=" + window.__crewId, RK = "https://playninjarift.com/api/crew_ranking_website.php", CA = "https://playninjarift.com/api/crew_ranking_castles_website.php";
  var tb = document.querySelector("tbody"), names = [], rws = tb.querySelectorAll("tr");
  for (var i = 0; i < rws.length; i++) names.push(rws[i].cells[1].textContent.trim());
  var autoSeconds = 60, autoEl = document.getElementById("auto-seconds"), searchEl = document.getElementById("search-input"), dotEl = document.getElementById("status-dot"), statusEl = document.getElementById("status-text");
  if (window.__hourlyCache && Object.keys(window.__hourlyCache).length > 0) { var _ts = ts(), _m = String(new Date().getMinutes()), _b = _m <= "1" ? "01" : (_m >= "31" && _m <= "32" ? "31" : _m); localStorage.setItem("nr_1h", JSON.stringify({b: _b, ts: _ts, rs: window.__hourlyCache})); }
  if (window.__30mCache && Object.keys(window.__30mCache).length > 0) { var _b30 = (new Date().getMinutes() <= 1 ? "01" : "31"); localStorage.setItem("nr_30m", JSON.stringify({b: _b30, ts: _ts, rs: window.__30mCache})); }
  function pad(n) { return n < 10 ? "0"+n : ""+n; }
  function ts() { var d = new Date(); return d.getFullYear()+"-"+pad(d.getMonth()+1)+"-"+pad(d.getDate())+" "+pad(d.getHours())+":"+pad(d.getMinutes())+":"+pad(d.getSeconds()); }
  function fj(u) { return fetch(u,{headers:{"Accept":"application/json"}}).then(function(r){return r.json();}).catch(function(){return null;}); }
  function dh(v) { return v > 0 ? '<span class="up">+'+v+"</span>" : v <= 0 ? '<span class="down">'+v+"</span>" : '<span class="na">N/A</span>'; }
  function blk30(m) { return m <= 1 ? "01" : (m >= 31 && m <= 32 ? "31" : null); }
  function blk1h(m) { return m <= 1 ? "01" : null; }
  function upd(d, rk) {
    autoSeconds = 60;
    var clan = null;
    if (rk) {
      if (Array.isArray(rk)) {
        for (var ci = 0; ci < rk.length; ci++) {
          if (rk[ci].crew_id === window.__crewId) { clan = rk[ci]; break; }
        }
      } else { clan = rk; }
    }
    var n = new Date(), nm = n.getMinutes(), ns = ts();
    var lm = {}; for (var i = 0; i < d.length; i++) { var _k = i < names.length ? names[i] : d[i].character_name; lm[_k] = {damage: d[i].member_damage, kills: d[i].boss_kills || 0}; }
    var n2r = {}, a = tb.querySelectorAll("tr"); for (var i = 0; i < a.length; i++) n2r[a[i].cells[1].textContent.trim()] = a[i];
    var c30 = null, c1h = null; try { c30 = JSON.parse(localStorage.getItem("nr_30m")); c1h = JSON.parse(localStorage.getItem("nr_1h")); } catch(e) {}
    if (!c1h && window.__hourlyCache) c1h = {rs: window.__hourlyCache, ts: ""};
    for (var i = 0; i < names.length; i++) {
      var name = names[i], md = lm[name]; if (!md) continue;
      var row = n2r[name]; if (!row) continue;
      var cel = row.cells;
      cel[2].textContent = md.kills;
      cel[5].textContent = md.damage;
      if (window.__phase !== 1) { cel[8].textContent = md.damage; }
    }
    for (var _n in lm) {
      if (names.indexOf(_n) === -1) {
        names.push(_n);
        var tr = document.createElement("tr");
        tr.className = "new-row";
        tr.innerHTML = '<td class="num"></td><td>' + _n + '</td><td class="num div-col">' + (lm[_n].kills||0) + '</td><td class="num ph1-diff"><span class="na">N/A</span></td><td class="num ph1-diff"><span class="na">N/A</span></td><td class="num">' + lm[_n].damage + '</td><td class="num ph1-diff"><span class="na">N/A</span></td><td class="num ph1-diff"><span class="na">N/A</span></td><td class="num div-col">' + (window.__phase !== 1 ? lm[_n].damage : '<span class="na">N/A</span>') + '</td><td class="num ph2-diff"><span class="na">N/A</span></td><td class="num ph2-diff"><span class="na">N/A</span></td><td class="num div-col"><span class="na">N/A</span></td><td class="num"><span class="na">N/A</span></td>';
        tb.appendChild(tr);
        if (searchEl && searchEl.value && _n.toLowerCase().indexOf(searchEl.value.toLowerCase()) === -1) tr.style.display = "none";
      }
    }
    a = tb.querySelectorAll("tr"); n2r = {};
    for (var i = 0; i < a.length; i++) n2r[a[i].cells[1].textContent.trim()] = a[i];
    for (var i = 0; i < names.length; i++) {
      var row = n2r[names[i]];
      if (row && !lm[names[i]]) row.className = "left-row";
    }
    if (clan) {
      var sv = document.querySelectorAll(".stats-col .stat-val");
      if (sv.length >= 2 && clan.crew_damage !== undefined) sv[1].textContent = Number(clan.crew_damage).toLocaleString();
    }
    var ft = document.querySelector(".footer");
    var st = document.getElementById("snapshot-ts"); if (st) st.textContent = ns;
    var rs = {}; for (var i = 0; i < d.length; i++) rs[d[i].character_name] = d[i].member_damage;
    var b30 = blk30(nm), b1h = blk1h(nm);
    if (b30 && (!c30 || c30.b !== b30)) localStorage.setItem("nr_30m", JSON.stringify({b: b30, ts: ns, rs: rs}));
    if (b1h && (!c1h || c1h.b !== b1h)) localStorage.setItem("nr_1h", JSON.stringify({b: b1h, ts: ns, rs: rs}));
    window.__defaultRows = tb.innerHTML;
    var sr = tb.querySelectorAll("tr");
    for (var ri = 0; ri < sr.length; ri++) sr[ri].cells[0].textContent = ri + 1;
  }
  function castleUpd(cdata) {
    var cb = document.getElementById("castle-bar");
    if (!cb || window.__phase !== 1) return;
    var cards = cb.querySelectorAll(".castle-card");
    var cache = null, bCache = null, tCache = null;
    try { cache = JSON.parse(localStorage.getItem("nr_castle_30m")); bCache = JSON.parse(localStorage.getItem("nr_castle_baseline")); tCache = JSON.parse(localStorage.getItem("nr_castle_trigger")); } catch(e) {}
    tCache = tCache || {};
    var newCache = {}, newBCache = {}, newTCache = {};
    for (var i = 0; i < cards.length; i++) {
      var card = cards[i];
      var csm = card.getAttribute("data-castle");
      if (!csm) { var cn = card.querySelector(".castle-name"); if (cn) csm = cn.textContent.trim(); else continue; }
      var entries = cdata[csm];
      if (!entries) continue;
      var our = null, rival = null;
      for (var ei = 0; ei < entries.length; ei++) {
        if (entries[ei].crew_id === window.__crewId) our = {rank: ei + 1, kills: entries[ei].boss_kills};
      }
      if (!our) continue;
      if (our.rank === 1) {
        if (entries.length < 2) continue;
        rival = {rank: 2, kills: entries[1].boss_kills, name: entries[1].crew_name};
      } else {
        var ri = our.rank - 2;
        if (ri >= 0 && ri < entries.length) rival = {rank: ri + 1, kills: entries[ri].boss_kills, name: entries[ri].crew_name};
      }
      if (!our || !rival) continue;
      var ourK = our.kills, rivalK = rival.kills;
      var pc = cache ? cache[csm] : null;
      var ourG = pc ? ourK - pc.our_kills : 0, rivalG = pc ? rivalK - pc.rival_kills : 0;
      newCache[csm] = {our_kills: ourK, rival_kills: rivalK, rival_name: rival.name, our_rank: our.rank};
      var bp = bCache ? bCache[csm] : null;
      newBCache[csm] = {our_kills: ourK, rival_kills: rivalK};
      var gStr = ourG > 0 ? "+" + ourG : "" + ourG;
      var rgStr = rivalG > 0 ? "+" + rivalG : "" + rivalG;
      var ourKEl = card.querySelector(".castle-kills.ours");
      var ourGEl = card.querySelector(".castle-gain.ours");
      var rivalKEl = card.querySelector(".castle-kills.rival");
      var rivalGEl = card.querySelector(".castle-gain.rival");
      var tagEl = card.querySelector(".castle-tag");
      var pillEl = card.querySelector(".castle-rank-pill");
      var rivalRankEl = card.querySelector(".castle-rival-rank");
      var rivalNameEl = card.querySelector(".castle-rival-name");
      if (ourKEl) ourKEl.textContent = ourK.toLocaleString();
      if (ourGEl) ourGEl.textContent = "(" + gStr + "/½h)";
      if (rivalKEl) rivalKEl.textContent = rivalK.toLocaleString();
      if (rivalGEl) rivalGEl.textContent = "(" + rgStr + "/½h)";
      if (pillEl) pillEl.textContent = "[#" + our.rank + "]";
      if (rivalRankEl) rivalRankEl.textContent = "#" + rival.rank;
      if (rivalNameEl) rivalNameEl.textContent = rival.name;
      var isLead = (our.rank === 1);
      var gapVal = isLead ? (ourK - rivalK) : (rivalK - ourK);
      var label = isLead ? "Lead " : "Need ";
      if (tagEl) tagEl.textContent = label + gapVal.toLocaleString();
      card.classList.remove("dangerous", "catching");
      tagEl.classList.remove("castle-tag-danger", "castle-tag-catch");
      var curGap = isLead ? (ourK - rivalK) : (rivalK - ourK);
      var trig = tCache[csm];
      var active = false;
      if (trig) {
        var ref = trig.ref_gap;
        var changePct = ref > 0 ? Math.round((curGap - ref) / ref * 100) : 0;
        if (trig.state === "dangerous" && changePct >= 15) {
          trig = null;
        } else if (trig.state === "catching" && changePct >= 15) {
          trig = null;
        } else {
          active = true;
          if (trig.state === "dangerous") { card.classList.add("dangerous"); tagEl.classList.add("castle-tag-danger"); }
          else { card.classList.add("catching"); tagEl.classList.add("castle-tag-catch"); }
        }
      }
      if (!trig) {
        var prevGap = 0;
        if (bp) { prevGap = isLead ? (bp.our_kills - bp.rival_kills) : (bp.rival_kills - bp.our_kills); }
        var pct = prevGap > 0 ? Math.round((prevGap - curGap) / prevGap * 100) : 0;
        if (isLead && pct >= 30) { trig = {state: "dangerous", ref_gap: curGap}; card.classList.add("dangerous"); tagEl.classList.add("castle-tag-danger"); }
        else if (!isLead && pct >= 30) { trig = {state: "catching", ref_gap: curGap}; card.classList.add("catching"); tagEl.classList.add("castle-tag-catch"); }
      }
      newTCache[csm] = trig || null;
    }
    var nm = new Date().getMinutes(), blk = nm <= 1 ? "01" : (nm >= 31 && nm <= 32 ? "31" : null);
    if (blk) localStorage.setItem("nr_castle_30m", JSON.stringify(newCache));
    localStorage.setItem("nr_castle_baseline", JSON.stringify(newBCache));
    localStorage.setItem("nr_castle_trigger", JSON.stringify(newTCache));
  }
  function refreshData() {
    if (dotEl) dotEl.className = "status-dot wait";
    if (statusEl) statusEl.textContent = "Loading...";
    Promise.all([fj(API), fj(RK), fj(CA).then(function(cdata) { if (cdata && window.__phase === 1) castleUpd(cdata); return cdata; }).catch(function() { return null; })]).then(function(r) {
      if (r[0] && r[0].members) {
        upd(r[0].members, r[1]);
        if (dotEl) dotEl.className = "status-dot ok";
        if (statusEl) statusEl.textContent = "Live";
      } else {
        if (dotEl) dotEl.className = "status-dot err";
        if (statusEl) statusEl.textContent = "Offline";
      }
    });
  }
  window.__refreshData = refreshData;
  refreshData();
  setInterval(refreshData, 60000);
  setInterval(function(){if(autoSeconds>0)autoSeconds--;if(autoEl)autoEl.textContent=autoSeconds;},1000);
  if(autoEl)autoEl.parentElement.addEventListener("click",function(){autoSeconds=60;refreshData();});
  if(searchEl)searchEl.addEventListener("input",function(){
    var q = this.value.toLowerCase(), r = tb.querySelectorAll("tr");
    for(var i=0;i<r.length;i++)r[i].style.display=r[i].cells[1].textContent.trim().toLowerCase().indexOf(q)>=0?"":"none";
  });
  // Click-to-copy (event delegation on tbody)
  tb.addEventListener("click", function(e) {
    var cell = e.target;
    while (cell && cell.tagName !== "TD") cell = cell.parentNode;
    if (!cell || cell.cellIndex !== 1) return;
    var name = cell.textContent.trim();
    navigator.clipboard.writeText(name).then(function() {
      var toast = document.createElement("span");
      toast.className = "copied-toast";
      toast.textContent = "Copied!";
      toast.style.opacity = "1";
      cell.style.position = "relative";
      cell.appendChild(toast);
      setTimeout(function() { toast.style.opacity = "0"; setTimeout(function() { toast.remove(); }, 300); }, 1500);
    }).catch(function() {});
  });
  // Reset sort
  var resetBtn = document.getElementById("reset-btn");
  if (resetBtn) resetBtn.addEventListener("click", function() {
    localStorage.removeItem("nr_sort");
    if (window.__resetSort) window.__resetSort();
  });
  // CSV export
  function csvDownload() {
    var rows = tb.querySelectorAll("tr"), csv = "Rank,Name,B.Kills,1/2H[K],Dmg,1/2H[D],Dmg(P2),1/2H[D](P2),Dmg,Kills\\n";
    for (var i = 0; i < rows.length; i++) {
      var cells = rows[i].cells, vals = [];
      for (var j = 0; j < cells.length; j++) {
        var v = cells[j].textContent.trim().replace(/"/g, '""');
        vals.push('"' + v + '"');
      }
      csv += vals.join(",") + "\\n";
    }
    var blob = new Blob([csv], {type: "text/csv;charset=utf-8"});
    var url = URL.createObjectURL(blob);
    var a = document.createElement("a");
    a.href = url; a.download = "SHAD0W NINJAS_reps.csv"; a.click();
    URL.revokeObjectURL(url);
  }
  var csvLink = document.getElementById("csv-link");
  if (csvLink) csvLink.addEventListener("click", csvDownload);
  // Updated ago
  function updateAgo() {
    var st = document.getElementById("snapshot-ts");
    var ua = document.getElementById("updated-ago");
    if (!st || !ua) return;
    var txt = st.textContent;
    var m = txt.match(/(\\d{4}-\\d{2}-\\d{2} \\d{2}:\\d{2}:\\d{2})/);
    if (!m) return;
    var snap = new Date(m[1].replace(" ", "T") + "+08:00").getTime();
    var now = new Date().getTime();
    var diff = Math.floor((now - snap) / 60000);
    ua.textContent = "Updated " + (diff < 1 ? "just now" : diff + "m ago");
  }
  updateAgo();
  setInterval(updateAgo, 30000);
})();
</script>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
{favicon_html}
<title>{crew_name} [Crew]</title>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{
    font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
    background: #0a0a0a;
    color: #e0e0e0;
    min-height: 100vh;
    display: flex;
    justify-content: center;
    padding: 32px 16px;
  }}
  .container {{
    max-width: 960px;
    width: 100%;
    box-shadow: 0 0 40px rgba(233, 69, 96, 0.06), 0 8px 32px rgba(0,0,0,0.5);
    border-radius: 16px;
    overflow: hidden;
  }}
  .header {{
    text-align: center;
    padding: 32px 24px 24px;
    background: linear-gradient(135deg, #121212 0%, #1a1a1a 50%, #101010 100%);
    position: relative;
    overflow: hidden;
  }}
  .header::before {{
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0; height: 3px;
    background: linear-gradient(90deg, #3b4db8, #5a6cc8, #3b4db8);
    background-size: 200% 100%;
    animation: shimmer 3s ease-in-out infinite;
  }}
  @keyframes shimmer {{ 0%,100% {{ background-position: 0% 50%; }} 50% {{ background-position: 100% 50%; }} }}
  .logo {{
    width: 132px; height: 132px;
    object-fit: contain;
    margin-bottom: 16px;
    filter: drop-shadow(0 0 16px rgba(233, 69, 96, 0.3));
  }}
  .header h1 {{
    font-size: 30px;
    font-weight: 700;
    color: #fff;
    letter-spacing: 0.5px;
    margin-bottom: 4px;
  }}
  .header .sub {{
    font-size: 17px;
    color: #888;
    display: flex;
    justify-content: center;
    gap: 16px;
    flex-wrap: wrap;
  }}
  .header .sub span {{ color: #aaa; }}
  .archive {{
    display: flex;
    gap: 6px;
    justify-content: center;
    padding: 14px 20px;
    background: #0f0f0f;
    border-bottom: 1px solid #1a1a2e;
    border-top: 1px solid #1a1a2e;
  }}
  .archive a {{
    color: #777;
    text-decoration: none;
    font-size: 12px;
    padding: 5px 14px;
    border-radius: 20px;
    border: 1px solid #1a1a2e;
    transition: 0.25s;
  }}
  .archive a:hover {{ border-color: #3b4db8; color: #fff; background: rgba(233, 69, 96, 0.08); }}
  .archive a.active {{ border-color: #3b4db8; color: #fff; background: #3b4db8; font-weight: 600; }}
  .table-wrap {{ overflow-x: auto; }}
  table {{
    width: 100%;
    min-width: 0;
    border-collapse: collapse;
    background: #0d0d0d;
  }}
  thead {{ position: sticky; top: 0; z-index: 1; }}
  th {{
    background: #111111;
    padding: 14px 18px;
    text-align: center;
    vertical-align: middle;
    font-size: 15px;
    line-height: 1;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: #4a6ad8;
    font-weight: 600;
    border-left: 2px solid transparent;
    box-sizing: border-box;
  }}
  td {{
    padding: 11px 18px;
    border-bottom: 1px solid #14141f;
    border-left: 2px solid transparent;
    font-size: 14px;
    color: #ccc;
    text-align: center;
    box-sizing: border-box;
  }}
  tr:nth-child(even) td {{ background: rgba(255,255,255,0.015); }}
  tr:hover td {{ background: rgba(233, 69, 96, 0.04); }}
  td.num {{ font-variant-numeric: tabular-nums; }}
  .up {{ color: #4caf50; }}
  .down {{ color: #f44336; }}
  .na {{ color: #555; }}
  .changes {{
    background: #0d0d0d;
    padding: 20px 24px;
    border-top: 1px solid #14141f;
  }}
  .changes-title {{
    font-size: 13px;
    color: #4a6ad8;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 14px;
    text-align: center;
  }}
  .changes-cols {{
    display: flex;
    gap: 24px;
    justify-content: center;
  }}
  .changes-col {{
    flex: 1;
    max-width: 320px;
  }}
  .changes-head {{
    font-size: 12px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    font-weight: 600;
    margin-bottom: 8px;
    text-align: center;
  }}
  .changes-head.left {{ color: #f44336; }}
  .changes-head.joined {{ color: #4caf50; }}
  .changes-col ul {{
    list-style: none;
    padding: 0;
    margin: 0;
    text-align: center;
  }}
  .changes-col li {{
    padding: 4px 0;
    font-size: 14px;
    color: #ccc;
    border-bottom: 1px solid #14141f;
  }}
  .changes-col li:last-child {{ border-bottom: none; }}
  .footer {{
    text-align: center;
    padding: 18px 20px;
    background: #0a0a0a;
    color: #444;
    font-size: 12px;
    border-top: 1px solid #12121e;
  }}
  .footer .ref {{ color: #555; font-size: 11px; margin-top: 2px; }}
  .footer a {{ color: #4a6ad8; text-decoration: none; }}
  .timer-bar {{
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 10px;
    padding: 12px 20px;
    background: #111111aa;
    backdrop-filter: blur(8px);
    -webkit-backdrop-filter: blur(8px);
    border-top: 1px solid #1a1a2e;
    font-size: 15px;
    flex-wrap: wrap;
    position: relative;
  }}
  .timer-left {{ display: flex; align-items: center; gap: 10px; }}
  .timer-season {{ color: #4a6ad8; font-weight: 700; letter-spacing: 0.5px; }}
  .timer-sep {{ color: #444; }}
  .timer-clock {{ display: flex; align-items: center; gap: 6px; }}
  .timer-digits {{ font-variant-numeric: tabular-nums; }}
  .timer-digits span:first-child {{ color: #4a6ad8; font-weight: 600; min-width: 28px; display: inline-block; text-align: center; }}
  .timer-unit {{ color: #888; font-size: 12px; margin-left: 1px; }}
  .timer-right {{ position: absolute; right: 20px; top: 50%; transform: translateY(-50%); cursor: pointer; font-size: 12px; color: #888; user-select: none; white-space: nowrap; }}
  .timer-right:hover {{ color: #4a6ad8; }}
  .stats-bar {{
    display: flex;
    flex-direction: column;
    align-items: center;
    gap: 8px;
    padding: 14px 20px;
    background: #111111aa;
    backdrop-filter: blur(8px);
    -webkit-backdrop-filter: blur(8px);
    border-top: 1px solid #1a1a2e;
  }}
  .stats-row {{
    display: flex;
    justify-content: center;
    gap: 48px;
    width: 100%;
  }}
  .stats-col {{
    display: flex;
    flex-direction: column;
    align-items: center;
    gap: 2px;
  }}
  .stat-label {{ color: #888; font-size: 12px; text-transform: uppercase; letter-spacing: 0.3px; }}
  .stat-val {{ color: #e0e0e0; font-size: 18px; font-weight: 700; font-variant-numeric: tabular-nums; }}
  #today-gain {{ color: #4caf50; }}
  .castle-bar {{
    display: flex;
    flex-direction: column;
    background: #111111;
    border-top: 2px solid #3b4db844;
    border-bottom: 1px solid #1a1a2e;
    padding: 0;
  }}
  .castle-header {{
    padding: 8px 16px;
    font-size: 15px;
    text-align: center;
    text-transform: uppercase;
    letter-spacing: 2px;
    color: #4a6ad8;
    font-weight: 600;
    background: rgba(0,0,0,0.25);
  }}
  .castle-grid {{
    display: flex;
    flex-wrap: wrap;
    justify-content: center;
    gap: 6px;
    padding: 6px;
  }}
  .castle-card {{
    flex: 0 0 calc((100% - 36px) / 7);
    background: #101010;
    border: 1px solid #1a1a2e;
    border-radius: 8px;
    padding: 6px 5px 6px;
    text-align: center;
    display: flex;
    flex-direction: column;
    align-items: center;
    gap: 1px;
    box-sizing: border-box;
  }}
  .castle-card.dangerous {{ border-color: #f4433666; }}
  .castle-card.catching {{ border-color: #4caf5066; }}
  .castle-emoji {{ font-size: 20px; line-height: 1; }}
  .castle-name {{ color: #ddd; font-size: 10px; font-weight: 600; }}
  .castle-rank-pill {{
    background: #3b4db822;
    color: #3b4db8;
    font-size: 9px;
    font-weight: 700;
    padding: 1px 6px;
    border-radius: 8px;
    margin: 1px 0 2px;
  }}
  .castle-rival-rank {{ color: #777; font-size: 9px; font-weight: 600; }}
  .castle-rival-name {{ color: #999; font-size: 9px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; max-width: 100%; }}
  .castle-kills {{ font-size: 13px; font-weight: 700; font-variant-numeric: tabular-nums; }}
  .castle-kills.ours {{ color: #fff; }}
  .castle-kills.rival {{ color: #bbb; }}
  .castle-gain {{ font-size: 9px; font-variant-numeric: tabular-nums; }}
  .castle-gain.ours {{ color: #4caf50; }}
  .castle-gain.rival {{ color: #777; }}
  .castle-div {{
    width: 80%;
    height: 1px;
    background: #1a1a2e;
    margin: 3px 0;
    flex-shrink: 0;
  }}
  .castle-tag {{
    margin-top: 3px;
    padding: 2px 8px;
    border-radius: 8px;
    font-size: 9px;
    font-weight: 600;
    color: #4a6ad8;
    background: #4a6ad818;
  }}
  .castle-tag-danger {{ color: #f44336; background: #f4433618; }}
  .castle-tag-catch {{ color: #4caf50; background: #4caf5018; }}
  @media (max-width: 600px) {{
    .castle-grid {{ padding: 6px; gap: 6px; }}
    .castle-card {{ flex: 0 0 100%; padding: 10px 8px 8px; gap: 3px; }}
    .castle-kills {{ font-size: 16px; }}
    .castle-emoji {{ font-size: 26px; }}
    .castle-name {{ font-size: 13px; }}
  }}
  th {{ cursor: pointer; user-select: none; }}
  th .sort-arrow {{ font-size: 11px; margin-left: 3px; }}
  @media (max-width: 600px) {{
    body {{ padding: 16px 8px; }}
    .header {{ padding: 24px 16px 20px; }}
    .header h1 {{ font-size: 22px; }}
    .header .sub {{ font-size: 14px; }}
    .logo {{ width: 96px; height: 96px; }}
    table {{ min-width: 520px; }}
    th, td {{ padding: 10px 10px; font-size: 12px; }}
    .changes {{ padding: 16px; }}
    .changes-cols {{ flex-direction: column; gap: 14px; }}
    .changes-col {{ max-width: 100%; }}
    .archive {{ flex-wrap: nowrap; overflow-x: auto; justify-content: flex-start; -webkit-overflow-scrolling: touch; scrollbar-width: none; }}
    .archive::-webkit-scrollbar {{ display: none; }}
    .archive a {{ flex-shrink: 0; }}
    .stats-bar {{ gap: 10px; padding: 12px 16px; }}
    .stat-val {{ font-size: 15px; }}
    .stats-row {{ flex-direction: column; gap: 6px; align-items: center; }}
    .stats-row + .stats-row {{ border-top: 1px solid #1a1a2e; padding-top: 10px; }}
  }}
  .live-bar {{
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 8px 20px;
    background: #0e0e0e;
    border-top: 1px solid #14141f;
    border-bottom: 1px solid #14141f;
    gap: 12px;
    flex-wrap: wrap;
  }}
  #search-input {{
    flex: 1;
    min-width: 160px;
    padding: 7px 12px;
    border-radius: 6px;
    border: 1px solid #1a1a2e;
    background: #111111;
    color: #e0e0e0;
    font-size: 13px;
    outline: none;
  }}
  #search-input:focus {{ border-color: #3b4db8; }}
  #search-input::placeholder {{ color: #555; }}
  .live-status {{ display: flex; align-items: center; gap: 6px; font-size: 12px; color: #555; white-space: nowrap; }}
  .status-dot {{ width: 8px; height: 8px; border-radius: 50%; }}
  .status-dot.ok {{ background: #4caf50; }}
  .status-dot.err {{ background: #f44336; }}
  .status-dot.wait {{ background: #888; animation: pulse 1.5s infinite; }}
  @keyframes pulse {{ 0%,100% {{ opacity: 1; }} 50% {{ opacity: 0.3; }} }}
  tr.new-row td {{ animation: fadeIn 0.5s ease; }}
  @keyframes fadeIn {{ from {{ opacity: 0; background: rgba(233,69,96,0.1); }} to {{ opacity: 1; background: transparent; }} }}
  tr.left-row td {{ opacity: 0.35; }}
  .goal-bar {{
    display: flex;
    flex-direction: column;
    gap: 4px;
    padding: 14px 20px;
    background: #0f0f0f;
    border-top: 1px solid #1a1a2e;
  }}
  .goal-track {{
    width: 100%;
    height: 16px;
    background: #14141f;
    border-radius: 8px;
    overflow: hidden;
  }}
  .goal-fill {{
    height: 100%;
    background: linear-gradient(90deg, #3b4db8, #5a6cc8);
    border-radius: 8px;
    transition: width 0.5s ease;
  }}
  .goal-info {{
    display: flex;
    justify-content: space-between;
    font-size: 12px;
    color: #888;
  }}
  .goal-info .goal-next {{ color: #4a6ad8; font-weight: 600; }}
  .goal-info .goal-num {{ color: #ccc; font-variant-numeric: tabular-nums; }}
  td:first-child {{ width: 28px; min-width: 28px; text-align: center; color: #666; font-size: 12px; }}
  .meg.divider {{ border-left: 2px solid #3b4db888; padding-left: 12px; }}
  .divider {{ border-left: 2px solid #3b4db888; }}
  .div-col {{ border-left: 2px solid #3b4db888; }}
  {hide_css}
  .action-btn {{ cursor: pointer; font-size: 12px; color: #888; padding: 4px 10px; border-radius: 4px; border: 1px solid #1a1a2e; background: #111111; user-select: none; white-space: nowrap; }}
  .action-btn:hover {{ border-color: #3b4db8; color: #3b4db8; }}
  .footer-updated {{ color: #555; font-size: 11px; margin: 2px 0; }}
  .footer-csv {{ margin-top: 8px; }}
  .footer-csv a {{ color: #4a6ad8; text-decoration: none; font-size: 12px; cursor: pointer; }}
  .footer-csv a:hover {{ text-decoration: underline; }}
  .copied-toast {{ position: absolute; background: #4a6ad8; color: #fff; font-size: 11px; padding: 2px 8px; border-radius: 4px; white-space: nowrap; pointer-events: none; opacity: 0; transition: opacity 0.3s; z-index: 10; }}
</style>
</head>
<body>
<div class="container">
  <div class="header">
    {logo_html}
    <h1>{crew_name}</h1>
    <div class="sub">
      <span>Crew ID: {CREW_ID}</span>
      <span>&middot;</span>
      <span>{member_count} members</span>
    </div>
  </div>
  {timer_html}
  {stats_html}
  {castle_html}
  <div class="live-bar">
    <div style="display:flex;align-items:center;gap:8px;flex:1;flex-wrap:wrap">
      <input type="text" id="search-input" placeholder="Search member...">
      <span class="action-btn" id="reset-btn">Reset</span>
    </div>
    <div class="live-status">
      <span class="status-dot wait" id="status-dot"></span>
      <span id="status-text">Idle</span>
    </div>
  </div>
  <div class="table-wrap">
  <table>
    <thead>
      <tr>
        <th rowspan="2" data-col="0"># <span class="sort-arrow"></span></th>
        <th rowspan="2" data-col="1">Name <span class="sort-arrow"></span></th>
        <th colspan="{p1_colspan}" class="meg">Phase 1</th>
        <th colspan="{p2_colspan}" class="meg divider">Phase 2</th>
        <th colspan="2" class="meg divider">DAILY</th>
      </tr>
      <tr>
        <th data-col="2" class="div-col">B.Kills <span class="sort-arrow"></span></th>
        <th data-col="3" class="ph1-diff">½H[K] <span class="sort-arrow"></span></th>
        <th data-col="4" class="ph1-diff">1H[K] <span class="sort-arrow"></span></th>
        <th data-col="5">Dmg <span class="sort-arrow"></span></th>
        <th data-col="6" class="ph1-diff">½H[D] <span class="sort-arrow"></span></th>
        <th data-col="7" class="ph1-diff">1H[D] <span class="sort-arrow"></span></th>
        <th data-col="8" class="div-col">Dmg <span class="sort-arrow"></span></th>
        <th data-col="9" class="ph2-diff">½H[D] <span class="sort-arrow"></span></th>
        <th data-col="10" class="ph2-diff">1H[D] <span class="sort-arrow"></span></th>
        <th data-col="11" class="div-col">Dmg <span class="sort-arrow"></span></th>
        <th data-col="12">Kills <span class="sort-arrow"></span></th>
      </tr>
    </thead>
    <tbody>{table_rows}</tbody>
  </table>
  </div>
  {changes_html}
  <div class="footer">
    <div class="footer-updated" id="updated-ago"></div>
    Snapshot: <span id="snapshot-ts">{ts_str}</span>
    <div class="ref">{ref_30m}{" &middot; " if ref_30m and (hourly_ref or daily_ref) else ""}{hourly_ref}{" &middot; " if hourly_ref and daily_ref else ""}{daily_ref}</div>
    <div class="footer-csv"><a id="csv-link">Download CSV</a></div>
  </div>
</div>
{script_html}
</body>
</html>"""

    index_path = "index.html"
    with open(index_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[{ts_str}] Saved {index_path}")

    if all_dates:
        archive_path = f"{date_str}.html"
        with open(archive_path, "w", encoding="utf-8") as f:
            f.write(html)
        print(f"[{ts_str}] Saved {archive_path}")

def save_snapshot(data):
    now = datetime.now(TARGET_TZ)
    is_daily = (now.hour == 13)
    sheet_name = now.strftime("%Y-%m-%d")

    prev_data, prev_timestamp = load_prev_from_xlsx(EXCEL_FILE, sheet_name)

    uniq = get_unique_names(data["members"])

    cache_1h = load_1h_cache()
    hourly_diffs = {}
    hourly_ts = ""
    if cache_1h and cache_1h.get("members"):
        hourly_ts = cache_1h.get("timestamp", "")
        for i, m in enumerate(data["members"]):
            uni_name = uniq[i][1]
            reps = m["member_damage"]
            memb = cache_1h["members"]
            if uni_name in memb:
                diff = reps - memb[uni_name]
                hourly_diffs[uni_name] = f"+{diff}" if diff > 0 else str(diff)
            else:
                hourly_diffs[uni_name] = "N/A"

    cache_30m = load_30m_cache()
    diffs_30m = {}
    ref_30m_ts = ""
    if cache_30m:
        ref_30m_ts = cache_30m.get("timestamp", "")
        for i, m in enumerate(data["members"]):
            uni_name = uniq[i][1]
            reps = m["member_damage"]
            if uni_name in cache_30m.get("members", {}):
                diff = reps - cache_30m["members"][uni_name]
                diffs_30m[uni_name] = f"+{diff}" if diff > 0 else str(diff)
            else:
                diffs_30m[uni_name] = "N/A"
    diff_30m_data = {"ts": ref_30m_ts, "diffs": diffs_30m}

    cache_30m_kills = load_30m_kills_cache()
    diffs_30m_kills = {}
    if cache_30m_kills:
        for i, m in enumerate(data["members"]):
            uni_name = uniq[i][1]
            kills = m.get("boss_kills", 0) or 0
            if uni_name in cache_30m_kills.get("members", {}):
                diff = kills - cache_30m_kills["members"][uni_name]
                diffs_30m_kills[uni_name] = f"+{diff}" if diff > 0 else str(diff)
            else:
                diffs_30m_kills[uni_name] = "N/A"

    cache_1h_kills = load_1h_kills_cache()
    diffs_1h_kills = {}
    if cache_1h_kills and cache_1h_kills.get("members"):
        for i, m in enumerate(data["members"]):
            uni_name = uniq[i][1]
            kills = m.get("boss_kills", 0) or 0
            if uni_name in cache_1h_kills["members"]:
                diff = kills - cache_1h_kills["members"][uni_name]
                diffs_1h_kills[uni_name] = f"+{diff}" if diff > 0 else str(diff)
            else:
                diffs_1h_kills[uni_name] = "N/A"

    try:
        season_info = fetch_season_info()
    except Exception:
        season_info = None

    try:
        ranking = fetch_crew_ranking()
        crew_damage = ranking.get("crew_damage", 0)
    except Exception:
        crew_damage = 0

    changes = load_changes()
    if cache_30m and "order" in cache_30m:
        now_ts = now.strftime("%Y-%m-%d %H:%M:%S")
        raw_prev = [re.sub(r' \(#\d+\)$', '', n) for n in cache_30m["order"]]
        raw_curr = [m["character_name"] for m in data["members"]]
        from collections import Counter
        prev_count, curr_count = Counter(raw_prev), Counter(raw_curr)
        existing_change_keys = set((c["type"], c["name"]) for c in changes)
        for name, cnt in prev_count.items():
            diff = cnt - curr_count.get(name, 0)
            for _ in range(diff):
                if ("left", name) not in existing_change_keys:
                    changes.append({"type": "left", "name": name, "detected_at": now_ts})
        for name, cnt in curr_count.items():
            diff = cnt - prev_count.get(name, 0)
            for _ in range(diff):
                if ("joined", name) not in existing_change_keys:
                    changes.append({"type": "joined", "name": name, "detected_at": now_ts})
        save_changes(changes)
  
    stats = {"today_gain": 0, "season_total": crew_damage}
    is_hourly_mark = (now.minute <= 1)
    is_30m_mark = (now.minute <= 1 or (now.minute >= 31 and now.minute <= 32))

    castle_stats = []
    if season_info and season_info.get("phase", 1) == 1:
        try:
            castles = fetch_castle_ranking()
            cache_30m_castle = load_30m_castle_cache()
            castle_cache_dict = {}
            if cache_30m_castle and "castles" in cache_30m_castle:
                castle_cache_dict = cache_30m_castle["castles"]
            new_castle_cache = {}
            for castle_name, entries in castles.items():
                our_idx = next((i for i, e in enumerate(entries) if e["crew_id"] == CREW_ID), None)
                if our_idx is None:
                    continue
                our_rank = our_idx + 1
                our_kills = entries[our_idx]["boss_kills"]
                if our_rank == 1:
                    if len(entries) < 2:
                        continue
                    rival_idx = 1
                else:
                    rival_idx = our_idx - 1
                rival = entries[rival_idx]
                rival_kills = rival["boss_kills"]
                gap = our_kills - rival_kills
                our_gain = 0
                rival_gain = 0
                if castle_name in castle_cache_dict:
                    pc = castle_cache_dict[castle_name]
                    our_gain = our_kills - pc.get("our_kills", our_kills)
                    rival_gain = rival_kills - pc.get("rival_kills", rival_kills)
                new_castle_cache[castle_name] = {"our_kills": our_kills, "rival_kills": rival_kills, "rival_name": rival["crew_name"], "our_rank": our_rank}
                castle_stats.append({
                    "name": castle_name, "rank": our_rank, "our_kills": our_kills,
                    "our_gain": our_gain, "rival_name": rival["crew_name"],
                    "rival_rank": rival_idx + 1, "rival_kills": rival_kills,
                    "rival_gain": rival_gain, "gap": gap
                })
            if is_30m_mark:
                save_30m_castle_cache(new_castle_cache, now.strftime("%Y-%m-%d %H:%M:%S"))
        except Exception as e:
            print(f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] Castle error: {e}")
            pass

    if season_info and season_info.get("phase", 1) == 2 and not os.path.exists(PHASE1_CACHE):
        phase1_data = {}
        for m in data["members"]:
            phase1_data[m["character_name"]] = {"damage": m["member_damage"], "kills": m.get("boss_kills", 0)}
        with open(PHASE1_CACHE, "w", encoding="utf-8") as f:
            json.dump(phase1_data, f)
        print(f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] Phase 1 snapshot saved")

    if is_daily:
        save_xlsx(data, prev_data, now, uniq)
        existing_html = [f.replace(".html", "") for f in os.listdir(".") if f.endswith(".html") and f[:4].isdigit() and f != "index.html"]
        all_dates = set(existing_html)
        all_dates.add(sheet_name)
        save_html(data, prev_data, prev_timestamp, hourly_diffs, hourly_ts, now, sorted(all_dates), show_changes=True, season_info=season_info, stats=stats, diff_30m=diff_30m_data, changes=changes, hourly_cache=cache_1h["members"] if cache_1h else {}, cache_30m=cache_30m, diffs_30m_kills=diffs_30m_kills, castle_stats=castle_stats, diffs_1h_kills=diffs_1h_kills)
    else:
        save_html(data, prev_data, prev_timestamp, hourly_diffs, hourly_ts, now, [], show_changes=False, season_info=season_info, stats=stats, diff_30m=diff_30m_data, changes=changes, hourly_cache=cache_1h["members"] if cache_1h else {}, cache_30m=cache_30m, diffs_30m_kills=diffs_30m_kills, castle_stats=castle_stats, diffs_1h_kills=diffs_1h_kills)

    save_30m_cache(data["members"], uniq)
    save_30m_kills_cache(data["members"], uniq)
    if cache_30m:
        save_1h_cache(cache_30m.get("members", {}), cache_30m.get("timestamp", ""))
    if cache_30m_kills:
        save_1h_kills_cache(cache_30m_kills.get("members", {}), cache_30m_kills.get("timestamp", ""))
    if is_hourly_mark:
        save_hourly_cache(data["members"], uniq, now)

    if season_info:
        end_dt = datetime(2026, 7, 19, 5, 0, 0, tzinfo=timezone.utc)
        if now >= end_dt:
            save_seasonal_xlsx(data["members"], season_info["season"])

    try:
        save_daily_history()
    except Exception as e:
        print(f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] History error: {e}")

def save_daily_history():
    if not os.path.exists(EXCEL_FILE):
        return
    wb = load_workbook(EXCEL_FILE)
    names = sorted([s.title for s in wb.worksheets if s.title != "Sheet1"])
    if len(names) < 2:
        return
    sheets_data = []
    for s in names:
        ws = wb[s]
        members = []
        for row in ws.iter_rows(min_row=4, max_col=2, values_only=True):
            name = str(row[0]).strip() if row[0] else ""
            if name and row[1] is not None:
                members.append((name, int(row[1])))
        sheets_data.append({"date": s, "members": members})
    css = """<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body { font-family: 'Segoe UI', system-ui, sans-serif; background: #0a0a0a; color: #e0e0e0; min-height: 100vh; display: flex; justify-content: center; padding: 32px 16px; }
  .container { max-width: 800px; width: 100%; box-shadow: 0 0 40px rgba(233,69,96,0.06), 0 8px 32px rgba(0,0,0,0.5); border-radius: 16px; overflow: hidden; }
  .header { text-align: center; padding: 28px 24px 20px; background: linear-gradient(135deg, #121212 0%, #1a1a1a 50%, #101010 100%); }
  .header h1 { font-size: 26px; font-weight: 700; color: #fff; margin-bottom: 4px; }
  .header .sub { font-size: 14px; color: #888; }
  .nav { display: flex; justify-content: space-between; padding: 12px 20px; background: #0f0f0f; border-top: 1px solid #1a1a2e; border-bottom: 1px solid #1a1a2e; }
  .nav a { color: #4a6ad8; text-decoration: none; font-size: 13px; }
  .nav a:hover { text-decoration: underline; }
  .nav .inactive { color: #444; pointer-events: none; }
  .table-wrap { overflow-x: auto; }
  table { width: 100%; border-collapse: collapse; background: #0d0d0d; }
  th { background: #111111; padding: 12px 16px; text-align: center; font-size: 13px; text-transform: uppercase; letter-spacing: 1px; color: #4a6ad8; font-weight: 600; }
  td { padding: 10px 16px; border-bottom: 1px solid #14141f; font-size: 13px; color: #ccc; text-align: center; }
  tr:nth-child(even) td { background: rgba(255,255,255,0.015); }
  .footer { text-align: center; padding: 16px 20px; background: #0a0a0a; color: #444; font-size: 12px; border-top: 1px solid #12121e; }
  .footer a { color: #4a6ad8; text-decoration: none; }
  .footer a:hover { text-decoration: underline; }
  .index-list { padding: 20px; background: #0d0d0d; }
  .index-list a { display: block; padding: 8px 14px; color: #ccc; text-decoration: none; font-size: 14px; border-bottom: 1px solid #14141f; }
  .index-list a:hover { background: rgba(233,69,96,0.04); color: #fff; }
  .index-list a:last-child { border-bottom: none; }
  .star-joined { color: #42a5f5; }
  .star-left { color: #f44336; }
</style>"""
    daily_pages = []
    for i in range(1, len(sheets_data)):
        prev, curr = sheets_data[i-1], sheets_data[i]
        date = curr["date"]
        dt = datetime.strptime(date, "%Y-%m-%d")
        day_name = dt.strftime("%a")
        prev_list, curr_list = prev["members"], curr["members"]
        max_len = max(len(prev_list), len(curr_list))
        gains = []
        for j in range(max_len):
            if j < len(prev_list) and j < len(curr_list):
                pname, prep = prev_list[j]
                cname, crep = curr_list[j]
                gains.append({"name": cname, "gain": crep - prep, "joined": False, "left": False})
            elif j < len(curr_list):
                cname, crep = curr_list[j]
                gains.append({"name": cname, "gain": None, "joined": True, "left": False})
            else:
                pname, prep = prev_list[j]
                gains.append({"name": pname, "gain": None, "joined": False, "left": True})
        gains.sort(key=lambda x: (x["gain"] is None, -(x["gain"] or 0)))
        daily_pages.append({"date": date, "day_name": day_name})
        rows_html = ""
        for idx, g in enumerate(gains, 1):
            star = ""
            if g["joined"]: star = '<span class="star-joined">&#9733;</span> '
            elif g["left"]: star = '<span class="star-left">&#9734;</span> '
            gain_str = f'+{g["gain"]:,}' if g["gain"] is not None else '<span class="star-left">N/A</span>'
            rows_html += f'<tr><td>{idx}</td><td>{star}{g["name"]}</td><td>{gain_str}</td></tr>\n'
        prev_link = f'<a href="history_{prev["date"]}.html">&larr; Previous</a>' if i > 1 else '<span class="inactive">&larr; Previous</span>'
        next_link = f'<a href="history_{sheets_data[i+1]["date"]}.html">Next &rarr;</a>' if i < len(sheets_data) - 1 else '<span class="inactive">Next &rarr;</span>'
        page_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Daily Reps · {date} ({day_name})</title>
{css}
</head>
<body>
<div class="container">
  <div class="header"><h1>SHAD0W NINJAS</h1><div class="sub">Daily Reps · {date} ({day_name})</div></div>
  <div class="nav">{prev_link}<a href="history.html">Index</a>{next_link}</div>
  <div class="table-wrap"><table><thead><tr><th>#</th><th>Name</th><th>Gain</th></tr></thead><tbody>{rows_html}</tbody></table></div>
  <div class="footer"><a href="index.html">&larr; Back to main page</a></div>
</div>
</body>
</html>"""
        with open(f"history_{date}.html", "w", encoding="utf-8") as f:
            f.write(page_html)
        print(f"[{datetime.now(TARGET_TZ).strftime('%Y-%m-%d %H:%M:%S')}] Saved history_{date}.html")
    index_rows = ""
    for dp in daily_pages:
        index_rows += f'<a href="history_{dp["date"]}.html">{dp["date"]} ({dp["day_name"]}) &rarr;</a>\n'
    index_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Daily Rep History · Season 61</title>
{css}
</head>
<body>
<div class="container">
  <div class="header"><h1>SHAD0W NINJAS</h1><div class="sub">Daily Rep History (Season 61)</div></div>
  <div class="index-list">{index_rows}</div>
  <div class="footer"><a href="index.html">&larr; Back to main page</a> &middot; <a href="https://github.com/nixervo/Shad0wNinjas-Crew">Source</a></div>
</div>
</body>
</html>"""
    with open("history.html", "w", encoding="utf-8") as f:
        f.write(index_html)
    print(f"[{datetime.now(TARGET_TZ).strftime('%Y-%m-%d %H:%M:%S')}] Saved history.html")

def run():
    print("Crew snapshot daemon started. Running every 30 minutes.")
    os.system('git config user.name "Crew-snapshot-bot"')
    os.system('git config user.email "bot@Crew-snapshot.local"')
    while True:
        now = datetime.now(TARGET_TZ)
        if now.minute < 1:
            target = now.replace(minute=1, second=0, microsecond=0)
        elif now.minute < 31:
            target = now.replace(minute=31, second=0, microsecond=0)
        else:
            target = (now + timedelta(hours=1)).replace(minute=1, second=0, microsecond=0)
        sleep_sec = (target - now).total_seconds()
        time.sleep(sleep_sec)
        try:
            data = fetch_crew()
            save_snapshot(data)
            ts = datetime.now(TARGET_TZ).strftime("%Y-%m-%d %H:%M:%S")
            os.system("git add -A")
            os.system(f'git commit -m "auto: snapshot {ts}" --allow-empty')
            os.system("git push")
        except Exception as e:
            print(f"[{datetime.now(TARGET_TZ).strftime('%Y-%m-%d %H:%M:%S')}] ERROR: {e}")

def fetch_once():
    data = fetch_crew()
    save_snapshot(data)

if __name__ == "__main__":
    if "--once" in sys.argv:
        fetch_once()
    else:
        run()
